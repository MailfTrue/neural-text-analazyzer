import random
import os

import openpyxl
from openpyxl.cell import Cell

from tqdm import tqdm
from .abstract import AbstractNeuralNetworkTrainer
from .mixins import EmotionNeuralNetworkTrainerMixin, RelevanceNeuralNetworkTrainerMixin


class ExcelNeuralnetworkTrainer(AbstractNeuralNetworkTrainer):
    def message_is_valid(self, message: list[Cell]):
        raise NotImplementedError()

    def get_message_text(self, message: list[Cell]):
        text = message[4].value
        return str(text) if text is not None else ''

    def get_spacy_label(self, message: list[Cell]):
        raise NotImplementedError()

    def load_training_data(self, data_directory: str = "./train", split: float = 0.8, limit: int = 0) -> tuple:
        messages = []
        file_names = list(filter(lambda x: x.endswith(".xlsx"), os.listdir(data_directory)))
        desc = "Loading training data, current file: {}"
        with tqdm(total=100) as pbar:
            for file_num, file_name in enumerate(file_names):
                pbar.set_description(desc.format(file_name))
                if not limit or len(messages) <= limit:
                    wb = openpyxl.load_workbook(f"{data_directory}/{file_name}")
                    ws = wb.active

                    pbar_iter_cost = (100 / len(file_names)) * (1 / (ws.max_row - 1))
                    for row in range(2, ws.max_row):
                        if self.message_is_valid(ws[row]):
                            text = self.get_message_text(ws[row])
                            spacy_label = self.get_spacy_label(ws[row])
                            messages.append((text, spacy_label))

                            if limit and len(messages) >= limit:
                                break
                        pbar.update(pbar_iter_cost)
        random.shuffle(messages)

        if limit:
            messages = messages[:limit]
        split = int(len(messages) * split)
        return messages[:split], messages[split:]


class EmotionExcelNeuralNetworkTrainer(ExcelNeuralnetworkTrainer, EmotionNeuralNetworkTrainerMixin):
    def get_spacy_label(self, message: list[Cell]):
        rate = str(message[8].value)
        return {"cats": {"pos": "1" == rate,
                         "neutral": "3" == rate,
                         "neg": "2" == rate}}

    def message_is_valid(self, message: list[Cell]):
        return self.get_message_text(message).strip() and str(message[8].value) in ["1", "2", "3"]


class RelevanceExcelNeuralNetworkTrainer(ExcelNeuralnetworkTrainer, RelevanceNeuralNetworkTrainerMixin):
    def get_spacy_label(self, message: list[Cell]):
        rate = str(message[8].value)
        return {"cats": {"not_relevance": "0" == rate,
                         "relevance": "0" != rate}}

    def message_is_valid(self, message: list[Cell]):
        return self.get_message_text(message).strip() and message[8].value
